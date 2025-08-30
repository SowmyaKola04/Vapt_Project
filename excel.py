import sys
import pandas as pd
import os
import csv
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

if len(sys.argv) != 2:
    print("Usage: python3 excel.py <path_to_csv>")
    sys.exit(1)

csv_path = sys.argv[1]
excel_path = csv_path.replace(".csv", ".xlsx")

# Clean and align CSV data
cleaned_rows = []
with open(csv_path, newline='', encoding='utf-8') as f:
    reader = csv.reader(f)
    headers = next(reader)
    while len(headers) < 4:
        headers.append("")
    for row in reader:
        while len(row) < 4:
            row.append("")
        cleaned_rows.append(row[:4])

# Create DataFrame
df = pd.DataFrame(cleaned_rows, columns=headers[:4])
df["Setting"] = df["Setting"].replace("", "Unknown Setting")

# Compliance calculation
total_checks = df.shape[0]
true_count = df["Status"].str.strip().str.lower().eq("true").sum()
compliance = round((true_count / total_checks) * 100, 2)

# Save to Excel
sheet_name = "Hardening Report"
df.to_excel(excel_path, index=False, sheet_name=sheet_name)

# Load workbook and worksheet
wb = load_workbook(excel_path)
ws = wb[sheet_name]

# Style definitions
bold_font = Font(bold=True)
green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
center_align = Alignment(horizontal="center")

# Apply bold to headers
for col in range(1, ws.max_column + 1):
    ws.cell(row=1, column=col).font = bold_font

# Status column color coding
status_col_index = None
for col in range(1, ws.max_column + 1):
    if str(ws.cell(row=1, column=col).value).strip().lower() == "status":
        status_col_index = col
        break

if status_col_index:
    for row in range(2, ws.max_row + 1):
        cell = ws.cell(row=row, column=status_col_index)
        val = str(cell.value).strip().lower()
        if val == "true":
            cell.fill = green_fill
        elif val == "false":
            cell.fill = red_fill

# Insert 2 blank rows after last data row
last_data_row = ws.max_row
for _ in range(2):
    ws.insert_rows(last_data_row + 1)

# Add Compliance and Risk rows with offset (2 columns right)
compliance_row = last_data_row + 3
ws.cell(row=compliance_row, column=3, value="Compliance (%)").font = bold_font
ws.cell(row=compliance_row, column=4, value=str(compliance)).font = bold_font
ws.cell(row=compliance_row, column=4).alignment = center_align

# Risk Status
risk_text = "✅ Device is Secured" if compliance >= 65 else "⚠️ Device is at RISK"
risk_fill = green_fill if compliance >= 65 else red_fill

risk_row = compliance_row + 1
ws.cell(row=risk_row, column=3, value="Risk Status").font = bold_font
risk_cell = ws.cell(row=risk_row, column=4, value=risk_text)
risk_cell.font = bold_font
risk_cell.fill = risk_fill
risk_cell.alignment = center_align

# Save and clean up
wb.save(excel_path)
os.remove(csv_path)
print(f"✅ Excel created: {excel_path}")
