import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from datetime import datetime
from pytz import timezone

# Base directory paths
base_dir = "/home/nipun/vapt_project/vapt_reports"
summary_dir = os.path.join(base_dir, "summary")
os.makedirs(summary_dir, exist_ok=True)

summary_data = []

for device in os.listdir(base_dir):
    device_path = os.path.join(base_dir, device)
    if not os.path.isdir(device_path) or device == "summary":
        continue

    timestamps = [
        ts for ts in os.listdir(device_path)
        if os.path.isdir(os.path.join(device_path, ts))
    ]
    if not timestamps:
        continue

    latest_ts = sorted(timestamps)[-1]
    report_dir = os.path.join(device_path, latest_ts)
    report_path = os.path.join(report_dir, "hardening_summary.xlsx")
    if not os.path.exists(report_path):
        continue

    try:
        df = pd.read_excel(report_path, sheet_name="Hardening Report")

        compliance_row = df[df["Recommendation"].str.strip().str.lower() == "compliance (%)"]
        status_row = df[df["Recommendation"].str.strip().str.lower() == "risk status"]

        compliance_value = None
        device_status = ""

        if not compliance_row.empty:
            reason_value = compliance_row.iloc[0]["Reason"]
            try:
                compliance_value = float(reason_value)
            except:
                compliance_value = None

        if not status_row.empty:
            device_status = status_row.iloc[0]["Reason"]

        if compliance_value is not None:
            summary_data.append({
                "Device": device,
                "Compliance (%)": compliance_value,
                "Status": device_status
            })

    except Exception as e:
        print(f"⚠️ Failed to read {report_path}: {e}")

if not summary_data:
    print("⚠️ No valid compliance data found. Exiting.")
    exit(1)

# Create DataFrame
summary_df = pd.DataFrame(summary_data)
summary_df.sort_values(by="Compliance (%)", ascending=False, inplace=True)

# Save with timestamp
timestamp = datetime.now(timezone("Asia/Kolkata")).strftime("%d-%m-%Y_%H-%M-%S")
output_path = os.path.join(summary_dir, f"compliance_summary_{timestamp}.xlsx")
summary_df.to_excel(output_path, index=False)

# Apply color formatting using openpyxl
wb = load_workbook(output_path)
ws = wb.active

green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # light green
red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")    # light red

for row in range(2, ws.max_row + 1):
    status_cell = ws.cell(row=row, column=3)  # Column C is "Status"
    status_value = str(status_cell.value).lower()

    if "secured" in status_value:
        status_cell.fill = green_fill
    elif "risk" in status_value:
        status_cell.fill = red_fill

wb.save(output_path)

print(f"✅ Summary report with color formatting saved to: {output_path}")
